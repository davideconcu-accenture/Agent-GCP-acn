"""
Script per generare file Excel BRB (Business Requirements Baseline)
per gli ETL di esempio
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os

def create_brb_vendite():
    """Crea il BRB per ETL Vendite"""
    wb = Workbook()
    
    # Rimuovi il foglio default
    wb.remove(wb.active)
    
    # === FOGLIO 1: Info Generali ===
    ws_info = wb.create_sheet("Info Generali")
    
    # Header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    # Info generali
    info_data = [
        ["Campo", "Valore"],
        ["Nome ETL", "ETL_VENDITE"],
        ["Versione", "1.0"],
        ["Data Creazione", "2025-01-15"],
        ["Owner", "Team Data Engineering"],
        ["Frequenza", "Giornaliera (ore 02:00)"],
        ["Obiettivo", "Caricare dati vendite giornaliere dalla sorgente al DWH"],
    ]
    
    for row_idx, row_data in enumerate(info_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_info.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
    
    # Auto-adjust columns
    for col in ws_info.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        ws_info.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2
    
    # === FOGLIO 2: Sorgenti ===
    ws_sorgenti = wb.create_sheet("Sorgenti")
    
    sorgenti_data = [
        ["Tabella", "Campo", "Tipo", "Descrizione", "Obbligatorio"],
        ["sorgente.vendite", "id_vendita", "PK", "Identificativo univoco vendita", "Sì"],
        ["sorgente.vendite", "data_vendita", "TIMESTAMP", "Data e ora transazione", "Sì"],
        ["sorgente.vendite", "id_cliente", "FK", "Riferimento cliente", "Sì"],
        ["sorgente.vendite", "id_prodotto", "FK", "Riferimento prodotto", "Sì"],
        ["sorgente.vendite", "quantita", "INT", "Unità vendute (>0)", "Sì"],
        ["sorgente.vendite", "prezzo_unitario", "DECIMAL", "Prezzo per unità (>0)", "Sì"],
        ["sorgente.vendite", "sconto", "INT", "Percentuale sconto (0-100)", "No"],
        ["sorgente.vendite", "id_negozio", "FK", "Punto vendita", "Sì"],
        ["sorgente.vendite", "stato", "VARCHAR", "Stato vendita", "Sì"],
        ["sorgente.clienti", "id_cliente", "PK", "Identificativo cliente", "Sì"],
        ["sorgente.clienti", "nome", "VARCHAR", "Nome cliente", "Sì"],
        ["sorgente.clienti", "cognome", "VARCHAR", "Cognome cliente", "Sì"],
        ["sorgente.clienti", "regione", "VARCHAR", "Area geografica", "Sì"],
        ["sorgente.prodotti", "id_prodotto", "PK", "Identificativo prodotto", "Sì"],
        ["sorgente.prodotti", "nome_prodotto", "VARCHAR", "Nome prodotto", "Sì"],
        ["sorgente.prodotti", "categoria", "VARCHAR", "Categoria merceologica", "Sì"],
    ]
    
    for row_idx, row_data in enumerate(sorgenti_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_sorgenti.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
    
    for col in ws_sorgenti.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        ws_sorgenti.column_dimensions[get_column_letter(col[0].column)].width = min(max_length + 2, 50)
    
    # === FOGLIO 3: Regole Business ===
    ws_regole = wb.create_sheet("Regole Business")
    
    regole_data = [
        ["ID Regola", "Descrizione", "Formula/Logica", "Criticità"],
        ["RB-001", "Periodo Elaborazione", "DATE(data_vendita) = CURRENT_DATE() - 1", "Alta"],
        ["RB-002", "Filtro Stato", "stato = 'CONFERMATA'", "Alta"],
        ["RB-003", "Join Obbligatori", "INNER JOIN con clienti e prodotti", "Alta"],
        ["RB-004", "Calcolo Importo Lordo", "quantita × prezzo_unitario", "Alta"],
        ["RB-005", "Calcolo Sconto", "(quantita × prezzo_unitario) × (sconto / 100)", "Alta"],
        ["RB-006", "Calcolo Importo Netto", "Importo Lordo - Importo Sconto", "Alta"],
        ["RB-007", "Nome Cliente", "CONCAT(nome, ' ', cognome)", "Media"],
        ["RB-008", "Validazione Quantità", "quantita > 0", "Alta"],
        ["RB-009", "Validazione Prezzo", "prezzo_unitario > 0", "Alta"],
        ["RB-010", "Validazione Sconto", "sconto BETWEEN 0 AND 100", "Media"],
    ]
    
    for row_idx, row_data in enumerate(regole_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_regole.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
    
    for col in ws_regole.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        ws_regole.column_dimensions[get_column_letter(col[0].column)].width = min(max_length + 2, 60)
    
    # === FOGLIO 4: Target ===
    ws_target = wb.create_sheet("Target")
    
    target_data = [
        ["Campo", "Tipo", "Descrizione", "Derivazione"],
        ["id_vendita", "PK", "Identificativo vendita", "sorgente.vendite.id_vendita"],
        ["data_vendita", "TIMESTAMP", "Data vendita", "sorgente.vendite.data_vendita"],
        ["id_cliente", "INT", "ID Cliente", "sorgente.vendite.id_cliente"],
        ["cliente_completo", "VARCHAR", "Nome completo cliente", "CONCAT(nome, ' ', cognome)"],
        ["regione", "VARCHAR", "Regione cliente", "sorgente.clienti.regione"],
        ["id_prodotto", "INT", "ID Prodotto", "sorgente.vendite.id_prodotto"],
        ["nome_prodotto", "VARCHAR", "Nome prodotto", "sorgente.prodotti.nome_prodotto"],
        ["categoria", "VARCHAR", "Categoria prodotto", "sorgente.prodotti.categoria"],
        ["quantita", "INT", "Quantità venduta", "sorgente.vendite.quantita"],
        ["prezzo_unitario", "DECIMAL", "Prezzo unitario", "sorgente.vendite.prezzo_unitario"],
        ["sconto", "INT", "Sconto %", "sorgente.vendite.sconto"],
        ["importo_lordo", "DECIMAL", "Importo lordo", "Calcolato"],
        ["importo_sconto", "DECIMAL", "Importo sconto", "Calcolato"],
        ["importo_netto", "DECIMAL", "Importo netto", "Calcolato"],
        ["id_negozio", "INT", "ID Negozio", "sorgente.vendite.id_negozio"],
        ["data_caricamento", "TIMESTAMP", "Data caricamento", "CURRENT_TIMESTAMP()"],
    ]
    
    for row_idx, row_data in enumerate(target_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_target.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
    
    for col in ws_target.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        ws_target.column_dimensions[get_column_letter(col[0].column)].width = min(max_length + 2, 50)
    
    # === FOGLIO 5: KPI e Soglie ===
    ws_kpi = wb.create_sheet("KPI e Soglie")
    
    kpi_data = [
        ["ID KPI", "Descrizione", "Metrica", "Soglia Minima", "Soglia Massima", "Azione"],
        ["KPI-001", "Volume Dati", "Record processati", "1.000", "50.000", "Alert se fuori range"],
        ["KPI-002", "Integrità Referenziale", "% record validi", "99.5%", "100%", "Alert se < 99.5%"],
        ["KPI-003", "Tempo Esecuzione", "Minuti", "-", "10", "Alert se > 10 min"],
        ["KPI-004", "Validazione Importi", "Record negativi", "0", "0", "Blocco caricamento"],
        ["KPI-005", "Record Scartati", "% scartati", "-", "1%", "Log dettagliato"],
    ]
    
    for row_idx, row_data in enumerate(kpi_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_kpi.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
    
    for col in ws_kpi.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        ws_kpi.column_dimensions[get_column_letter(col[0].column)].width = min(max_length + 2, 40)
    
    return wb

def create_brb_ordini():
    """Crea il BRB per ETL Ordini"""
    wb = Workbook()
    wb.remove(wb.active)
    
    # === FOGLIO 1: Info Generali ===
    ws_info = wb.create_sheet("Info Generali")
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    info_data = [
        ["Campo", "Valore"],
        ["Nome ETL", "ETL_ORDINI"],
        ["Versione", "1.0"],
        ["Data Creazione", "2025-01-15"],
        ["Owner", "Team Data Engineering"],
        ["Frequenza", "Giornaliera (ore 03:00)"],
        ["Obiettivo", "Caricare e aggregare dati ordini con righe ordine"],
    ]
    
    for row_idx, row_data in enumerate(info_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_info.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
    
    for col in ws_info.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        ws_info.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2
    
    # === FOGLIO 2: Regole Business ===
    ws_regole = wb.create_sheet("Regole Business")
    
    regole_data = [
        ["ID Regola", "Descrizione", "Formula/Logica", "Criticità"],
        ["RB-001", "Periodo Elaborazione", "DATE(data_ordine) >= CURRENT_DATE() - 1", "Alta"],
        ["RB-002", "Join Cliente", "LEFT JOIN con clienti", "Media"],
        ["RB-003", "Aggregazione Righe", "COUNT, SUM su righe_ordine per ordine", "Alta"],
        ["RB-004", "Calcolo Importo Totale", "SUM(prezzo_unitario * quantita)", "Alta"],
        ["RB-005", "Flag Completato", "1 se stato='COMPLETATO', 0 altrimenti", "Media"],
        ["RB-006", "Merge Target", "MERGE su id_ordine (UPDATE o INSERT)", "Alta"],
    ]
    
    for row_idx, row_data in enumerate(regole_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_regole.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
    
    for col in ws_regole.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        ws_regole.column_dimensions[get_column_letter(col[0].column)].width = min(max_length + 2, 60)
    
    # === FOGLIO 3: KPI ===
    ws_kpi = wb.create_sheet("KPI e Soglie")
    
    kpi_data = [
        ["ID KPI", "Descrizione", "Metrica", "Soglia", "Azione"],
        ["KPI-001", "Volume Ordini", "Ordini processati", "> 100/giorno", "Alert se < 50"],
        ["KPI-002", "Tempo Esecuzione", "Minuti", "< 8 minuti", "Alert se > 15"],
        ["KPI-003", "Ordini senza Righe", "Count", "0", "Log e investigazione"],
    ]
    
    for row_idx, row_data in enumerate(kpi_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_kpi.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
    
    for col in ws_kpi.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        ws_kpi.column_dimensions[get_column_letter(col[0].column)].width = min(max_length + 2, 40)
    
    return wb

def main():
    """Genera i file BRB Excel"""
    
    # Crea directory se non esiste
    os.makedirs("data/BRB/etl_vendite", exist_ok=True)
    os.makedirs("data/BRB/etl_ordini", exist_ok=True)
    
    # Genera BRB Vendite
    print("Generando BRB per ETL Vendite...")
    wb_vendite = create_brb_vendite()
    wb_vendite.save("data/BRB/etl_vendite/brb_vendite.xlsx")
    print("✓ Creato: data/BRB/etl_vendite/brb_vendite.xlsx")
    
    # Genera BRB Ordini
    print("Generando BRB per ETL Ordini...")
    wb_ordini = create_brb_ordini()
    wb_ordini.save("data/BRB/etl_ordini/brb_ordini.xlsx")
    print("✓ Creato: data/BRB/etl_ordini/brb_ordini.xlsx")
    
    print("\n✅ File BRB generati con successo!")

if __name__ == "__main__":
    main()