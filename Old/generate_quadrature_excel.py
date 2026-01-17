"""
Script per generare file Excel di Quadratura
Simula l'output di un quadratore Spark che confronta vecchio vs nuovo workflow
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
from datetime import datetime

def apply_header_style(ws, row=1):
    """Applica stile agli header"""
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in ws[row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

def create_quadratura_vendite():
    """Crea Excel di quadratura per ETL Vendite"""
    wb = Workbook()
    wb.remove(wb.active)
    
    # === SHEET 1: Summary ===
    ws_summary = wb.create_sheet("Summary")
    
    summary_data = [
        ["Metrica", "Valore"],
        ["Data Esecuzione", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["ETL", "ETL_VENDITE"],
        ["", ""],
        ["TOTALE RECORD VECCHIO WORKFLOW", "12.458"],
        ["TOTALE RECORD NUOVO WORKFLOW", "12.450"],
        ["", ""],
        ["RECORD IN MATCH", "12.398"],
        ["RECORD SOLO IN VECCHIO", "60"],
        ["RECORD SOLO IN NUOVO", "52"],
        ["", ""],
        ["PERCENTUALE MATCH", "99.52%"],
        ["", ""],
        ["CAMPI DIFFERENTI", "142"],
        ["RECORD CON DIFFERENZE", "89"],
        ["", ""],
        ["ESITO", "⚠️ ATTENZIONE - Squadrature rilevate"],
    ]
    
    for row_idx, row_data in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
            elif value in ["⚠️ ATTENZIONE - Squadrature rilevate"]:
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                cell.font = Font(bold=True)
    
    ws_summary.column_dimensions['A'].width = 40
    ws_summary.column_dimensions['B'].width = 40
    
    # === SHEET 2: Match Details ===
    ws_match = wb.create_sheet("Match")
    
    match_data = [
        ["ID", "Totale Record Match", "Percentuale"],
        ["id_vendita", "12.398", "99.52%"],
    ]
    
    for row_idx, row_data in enumerate(match_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            ws_match.cell(row=row_idx, column=col_idx, value=value)
    
    apply_header_style(ws_match)
    
    for col in ws_match.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        ws_match.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2
    
    # === SHEET 3: Squadrature ===
    ws_squad = wb.create_sheet("Squadrature")
    
    squad_data = [
        ["Tipo", "ID Vendita", "Descrizione"],
        ["SOLO_VECCHIO", "V_10234", "Record presente solo nel vecchio workflow"],
        ["SOLO_VECCHIO", "V_10567", "Record presente solo nel vecchio workflow"],
        ["SOLO_VECCHIO", "V_11023", "Record presente solo nel vecchio workflow"],
        ["SOLO_NUOVO", "V_20145", "Record presente solo nel nuovo workflow"],
        ["SOLO_NUOVO", "V_20678", "Record presente solo nel nuovo workflow"],
    ]
    
    for row_idx, row_data in enumerate(squad_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_squad.cell(row=row_idx, column=col_idx, value=value)
            if row_idx > 1:
                if value == "SOLO_VECCHIO":
                    cell.fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
                elif value == "SOLO_NUOVO":
                    cell.fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    
    apply_header_style(ws_squad)
    
    for col in ws_squad.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        ws_squad.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2
    
    # === SHEET 4: Campi Differenti ===
    ws_diff = wb.create_sheet("Campi Differenti")
    
    diff_data = [
        ["ID Vendita", "Campo", "Valore Vecchio", "Valore Nuovo", "Tipo Differenza"],
        ["V_00123", "importo_netto", "150.00", "148.50", "Calcolo sconto differente"],
        ["V_00123", "importo_sconto", "15.00", "16.50", "Calcolo sconto differente"],
        ["V_00456", "cliente_completo", "Mario Rossi", "Mario  Rossi", "Spazio doppio nel nome"],
        ["V_00789", "importo_lordo", "200.00", "200.00", "Match (arrotondamento)"],
        ["V_00789", "importo_netto", "180.00", "179.99", "Differenza arrotondamento"],
        ["V_01234", "regione", "Lombardia", "LOMBARDIA", "Case sensitive"],
        ["V_01567", "quantita", "5", "5.0", "Formato numerico diverso"],
        ["V_02345", "data_vendita", "2025-01-15 10:30:00", "2025-01-15 10:30:00.000", "Precisione timestamp"],
    ]
    
    for row_idx, row_data in enumerate(diff_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_diff.cell(row=row_idx, column=col_idx, value=value)
            if row_idx > 1 and col_idx == 5:
                if "arrotondamento" in str(value).lower():
                    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                elif "case sensitive" in str(value).lower() or "spazio" in str(value).lower():
                    cell.fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
                elif "calcolo" in str(value).lower():
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    apply_header_style(ws_diff)
    
    for col in ws_diff.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        ws_diff.column_dimensions[get_column_letter(col[0].column)].width = min(max_length + 2, 50)
    
    return wb

def create_quadratura_ordini():
    """Crea Excel di quadratura per ETL Ordini"""
    wb = Workbook()
    wb.remove(wb.active)
    
    # === SHEET 1: Summary ===
    ws_summary = wb.create_sheet("Summary")
    
    summary_data = [
        ["Metrica", "Valore"],
        ["Data Esecuzione", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["ETL", "ETL_ORDINI"],
        ["", ""],
        ["TOTALE RECORD VECCHIO WORKFLOW", "3.567"],
        ["TOTALE RECORD NUOVO WORKFLOW", "3.567"],
        ["", ""],
        ["RECORD IN MATCH", "3.545"],
        ["RECORD SOLO IN VECCHIO", "22"],
        ["RECORD SOLO IN NUOVO", "22"],
        ["", ""],
        ["PERCENTUALE MATCH", "99.38%"],
        ["", ""],
        ["CAMPI DIFFERENTI", "67"],
        ["RECORD CON DIFFERENZE", "45"],
        ["", ""],
        ["ESITO", "⚠️ ATTENZIONE - Verificare aggregazioni"],
    ]
    
    for row_idx, row_data in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
            elif "⚠️" in str(value):
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                cell.font = Font(bold=True)
    
    ws_summary.column_dimensions['A'].width = 40
    ws_summary.column_dimensions['B'].width = 40
    
    # === SHEET 2: Campi Differenti ===
    ws_diff = wb.create_sheet("Campi Differenti")
    
    diff_data = [
        ["ID Ordine", "Campo", "Valore Vecchio", "Valore Nuovo", "Tipo Differenza"],
        ["O_10001", "importo_totale", "450.00", "449.95", "Aggregazione SUM con arrotondamento"],
        ["O_10002", "num_prodotti", "3", "3.0", "Formato COUNT diverso"],
        ["O_10003", "lista_prodotti", "Prod A, Prod B", "Prod A,Prod B", "Separatore STRING_AGG"],
        ["O_10004", "quantita_totale", "15", "15.00", "Tipo dato diverso"],
        ["O_10005", "flag_completato", "1", "TRUE", "Formato booleano diverso"],
    ]
    
    for row_idx, row_data in enumerate(diff_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_diff.cell(row=row_idx, column=col_idx, value=value)
            if row_idx > 1 and col_idx == 5:
                if "aggregazione" in str(value).lower():
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                elif "formato" in str(value).lower():
                    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    apply_header_style(ws_diff)
    
    for col in ws_diff.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        ws_diff.column_dimensions[get_column_letter(col[0].column)].width = min(max_length + 2, 50)
    
    # === SHEET 3: Raccomandazioni ===
    ws_racc = wb.create_sheet("Raccomandazioni")
    
    racc_data = [
        ["Priorità", "Problema", "Raccomandazione"],
        ["ALTA", "Aggregazione SUM con differenze", "Verificare arrotondamenti nelle somme. Usare ROUND() consistente."],
        ["MEDIA", "Formato COUNT diverso", "Standardizzare cast a INT per funzioni COUNT."],
        ["BASSA", "Separatore STRING_AGG", "Uniformare separatori nelle concatenazioni (spazio dopo virgola)."],
        ["MEDIA", "Flag booleano inconsistente", "Usare CASE WHEN per garantire formato numerico 0/1."],
    ]
    
    for row_idx, row_data in enumerate(racc_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_racc.cell(row=row_idx, column=col_idx, value=value)
            if row_idx > 1 and col_idx == 1:
                if value == "ALTA":
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                elif value == "MEDIA":
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                elif value == "BASSA":
                    cell.fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    
    apply_header_style(ws_racc)
    
    for col in ws_racc.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        ws_racc.column_dimensions[get_column_letter(col[0].column)].width = min(max_length + 2, 60)
    
    return wb

def main():
    """Genera i file di quadratura Excel"""
    
    # Crea directory se non esiste
    os.makedirs("data/Quadrature/etl_vendite", exist_ok=True)
    os.makedirs("data/Quadrature/etl_ordini", exist_ok=True)
    
    # Genera Quadratura Vendite
    print("Generando quadratura per ETL Vendite...")
    wb_vendite = create_quadratura_vendite()
    wb_vendite.save("data/Quadrature/etl_vendite/quadratura_vendite_2025-01-15.xlsx")
    print("✓ Creato: data/Quadrature/etl_vendite/quadratura_vendite_2025-01-15.xlsx")
    
    # Genera Quadratura Ordini
    print("Generando quadratura per ETL Ordini...")
    wb_ordini = create_quadratura_ordini()
    wb_ordini.save("data/Quadrature/etl_ordini/quadratura_ordini_2025-01-15.xlsx")
    print("✓ Creato: data/Quadrature/etl_ordini/quadratura_ordini_2025-01-15.xlsx")
    
    print("\n✅ File di quadratura generati con successo!")

if __name__ == "__main__":
    main()